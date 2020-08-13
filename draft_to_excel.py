import requests
import pandas as pd
from openpyxl import Workbook, load_workbook
'''
TODO
- create input for username
- create input for year of draft
- create input for new draft file
'''


'''
function: get_all_user_drafts(username, year)
parameters:
  username - Sleeper.app username. Required for calls to Sleeper API for gathering draft data.
  year - Year from which to pull draft(s) data. Required for calls to Sleeper API for gathering draft data.
description:
  Given a username and draft year, pull all draft data from Sleeper.app API for that user and year.
  For each of the completed drafts of that year, create a diction
returns:
  List (drafts) of Lists (draft picks) of Dict (individual pick)
'''
def get_all_user_drafts(user, year):
  
  raw_user_drafts = requests.get(url="https://api.sleeper.app/v1/user/" + user['user_id'] + "/drafts/nfl/" + str(year))
  raw_user_drafts = raw_user_drafts.json()
  user_drafts = []

  for draft in raw_user_drafts:
    if draft['status'] == "complete":
      
      raw_curr_draft = requests.get(url="https://api.sleeper.app/v1/draft/" + draft['draft_id'] + "/picks")
      curr_draft = raw_curr_draft.json()
      
      user_picks = []
      for pick in curr_draft:
        if pick['picked_by'] == user['user_id']:
            
            curr_user_pick = {}

            curr_user_pick['round'] = pick['round']
            curr_user_pick['pick_no'] = pick['pick_no']
            curr_user_pick['name'] = pick['metadata']['first_name'] + " " + pick['metadata']['last_name']
            curr_user_pick['team'] = pick['metadata']['team']
            curr_user_pick['position'] = pick['metadata']['position']

            user_picks.append(curr_user_pick)
      user_draft = {
                    'draft_id' : draft['draft_id'],
                    'draft_picks' : user_picks
                  }
      user_drafts.append(user_draft)
  
  return user_drafts

def main():
  
  user = None
  while user == None:
    print('Enter your Sleeper.app username: ')
    username = input()
    req = requests.get(url="https://api.sleeper.app/v1/user/" + str(username))
    user = req.json()
    if user == None:
      print("Invaid username!")
  
  path = './sleeper_drafts_2020.xlsx'
  try:
    book = load_workbook(path)
  except FileNotFoundError:
    print("File does not exist...creating new draft file...")
    wb = Workbook()
    wb.save(path)
    book = load_workbook(path)
  
  writer = pd.ExcelWriter(path, engine='openpyxl')
  writer.book = book

  print("Gathering draft data...")
  user_drafts = get_all_user_drafts(user, 2020)

  print("Writing draft data to '" + path + "'...")
  for draft in user_drafts:
    if draft['draft_id'] not in book.sheetnames:
        df = pd.DataFrame(draft['draft_picks'])
        df = df[['pick_no','round', 'name', 'team', 'position']]
        df.to_excel(writer, index=False, sheet_name=draft['draft_id'])
        book[str(draft['draft_id'])].column_dimensions['c'].width = 30
    
  writer.save()
  writer.close()
  print("Complete")
main()